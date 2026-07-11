"""Read XLSX workbook data for the spreadsheet viewer API."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.ingestion.xlsx_extract import is_sheet_visible, rows_from_range_numbered

logger = logging.getLogger(__name__)


def find_xlsx_path(user_id: int, doc_id: str) -> Path | None:
    from app.ingestion.pipeline import document_upload_dir

    dest_dir = document_upload_dir(user_id, doc_id)
    if not dest_dir.exists():
        return None
    matches = sorted(dest_dir.glob("*.xlsx"))
    return matches[0] if matches else None


def list_workbook_sheets(xlsx_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    visible_index = 0
    try:
        for worksheet in workbook.worksheets:
            if not is_sheet_visible(worksheet):
                continue
            visible_index += 1
            sheets.append(
                {
                    "name": worksheet.title,
                    "index": visible_index,
                    "row_count": int(worksheet.max_row or 0),
                    "col_count": int(worksheet.max_column or 0),
                }
            )
    finally:
        workbook.close()
    return sheets


def read_sheet_grid(
    xlsx_path: Path,
    sheet_name: str,
) -> dict[str, Any]:
    workbook = load_workbook(xlsx_path, read_only=False, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(sheet_name)
        worksheet = workbook[sheet_name]
        if not is_sheet_visible(worksheet):
            raise KeyError(sheet_name)

        visible_index = 0
        for ws in workbook.worksheets:
            if not is_sheet_visible(ws):
                continue
            visible_index += 1
            if ws.title == sheet_name:
                break

        min_row = 1
        max_row = int(worksheet.max_row or 1)
        min_col = 1
        max_col = int(worksheet.max_column or 1)
        numbered = rows_from_range_numbered(worksheet, min_row, min_col, max_row, max_col)
        row_numbers = [sheet_row for sheet_row, _ in numbered]
        logger.info(
            "XLSX_GRID sheet=%s rows=%s row_numbers_sample=%s tail=%s",
            sheet_name,
            len(numbered),
            row_numbers[:5],
            row_numbers[-3:] if row_numbers else [],
        )

        return {
            "name": sheet_name,
            "index": visible_index,
            "rows": [cells for _, cells in numbered],
            "row_numbers": [sheet_row for sheet_row, _ in numbered],
            "row_count": int(worksheet.max_row or 0),
            "col_count": int(worksheet.max_column or 0),
        }
    finally:
        workbook.close()


def resolve_sheet_by_index(xlsx_path: Path, sheet_index: int) -> str | None:
    sheets = list_workbook_sheets(xlsx_path)
    for sheet in sheets:
        if sheet["index"] == sheet_index:
            return sheet["name"]
    return None
