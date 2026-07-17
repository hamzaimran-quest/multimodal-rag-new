"""Build enriched XLSX chunks using validated workbook schema (Approach C)."""

from __future__ import annotations

import logging
from typing import Any

from app.ingestion.models import ExtractedChunk
from app.ingestion.xlsx_extract import (
    MIN_TABLE_WORDS,
    _chunks_from_row_bands,
    _chunks_from_tables,
)
from app.ingestion.xlsx_entity_keys import annotate_chunk_entity_keys
from app.ingestion.xlsx_schema import (
    ValidatedCluster,
    ValidatedSatellite,
    ValidatedWorkbookSchema,
    WorkbookSchemaRecognitionError,
    detect_and_validate_workbook_schema,
)
from app.ingestion.xlsx_serialize import resolve_row_band_size, rows_to_slim_values_text
from app.ingestion.xlsx_workbook import (
    SheetData,
    WorkbookData,
    load_workbook_data,
    resolve_header_index,
)

logger = logging.getLogger(__name__)

SUMMARY_PREVIEW_COUNT = 3


def _field_label(satellite: ValidatedSatellite, column_name: str) -> str:
    return f"{satellite.sheet_name}.{column_name}"


def _build_satellite_lookups(
    workbook: WorkbookData,
    cluster: ValidatedCluster,
) -> dict[str, dict[str, Any]]:
    """Per-satellite lookup tables keyed by primary key value."""
    primary_sheet = workbook.sheet_by_name[cluster.primary_sheet]
    lookups: dict[str, dict[str, Any]] = {}

    for satellite in cluster.satellites:
        sheet = workbook.sheet_by_name[satellite.sheet_name]
        one_to_one: dict[str, dict[str, str]] = {}
        one_to_many: dict[str, list[dict[str, Any]]] = {}

        for row_num, cells in sheet.rows:
            key_index = satellite.key_col_index
            if key_index >= len(cells):
                continue
            key_value = cells[key_index].strip()
            if not key_value:
                continue

            payload: dict[str, str] = {}
            origins: list[dict[str, Any]] = []
            for col_index in satellite.payload_col_indices:
                if col_index >= len(cells):
                    continue
                header = sheet.headers[col_index]
                payload[header] = cells[col_index]
                origins.append(
                    {
                        "field": _field_label(satellite, header),
                        "source_sheet": sheet.name,
                        "source_row": row_num,
                        "source_col": col_index + 1,
                    }
                )

            entry = {"values": payload, "origins": origins, "source_row": row_num}
            if satellite.cardinality == "one_to_one":
                one_to_one[key_value] = entry
            else:
                one_to_many.setdefault(key_value, []).append(entry)

        lookups[satellite.sheet_name] = {
            "satellite": satellite,
            "one_to_one": one_to_one,
            "one_to_many": one_to_many,
        }
    return lookups


def _summarize_many(values: list[str]) -> str:
    cleaned = [value for value in values if value]
    if not cleaned:
        return ""
    preview = cleaned[:SUMMARY_PREVIEW_COUNT]
    text = ", ".join(preview)
    remaining = len(cleaned) - len(preview)
    if remaining > 0:
        text = f"{text} (+{remaining} more)"
    return text


def _enrich_primary_row(
    primary_sheet: SheetData,
    row_num: int,
    cells: list[str],
    *,
    cluster: ValidatedCluster,
    lookups: dict[str, dict[str, Any]],
    enrichment_columns: list[tuple[ValidatedSatellite, str, str]],
) -> tuple[list[str], list[dict[str, Any]], str | None]:
    key_index = cluster.primary_key_col_index
    key_value = cells[key_index].strip() if key_index < len(cells) else ""
    extra_values: list[str] = []
    origins: list[dict[str, Any]] = [
        {
            "field": cluster.primary_key_column,
            "source_sheet": primary_sheet.name,
            "source_row": row_num,
            "source_col": key_index + 1,
        }
    ]

    for satellite, payload_header, mode in enrichment_columns:
        satellite_lookup = lookups[satellite.sheet_name]
        if mode == "scalar":
            entry = satellite_lookup["one_to_one"].get(key_value)
            if entry:
                value = entry["values"].get(payload_header, "")
                extra_values.append(value)
                for origin in entry["origins"]:
                    if origin["field"].endswith(payload_header) or origin["field"].endswith(
                        f"{satellite.sheet_name}.{payload_header}"
                    ):
                        origins.append(origin)
            else:
                extra_values.append("")
        else:
            entries = satellite_lookup["one_to_many"].get(key_value, [])
            values = [entry["values"].get(payload_header, "") for entry in entries]
            extra_values.append(_summarize_many(values))
            if entries:
                origins.append(
                    {
                        "field": f"{satellite.sheet_name}.{payload_header}",
                        "source_sheet": satellite.sheet_name,
                        "source_row": entries[0]["source_row"],
                        "source_col": satellite.key_col_index + 1,
                    }
                )

    return extra_values, origins, key_value or None


def _enrichment_columns_for_cluster(cluster: ValidatedCluster) -> list[tuple[ValidatedSatellite, str, str]]:
    columns: list[tuple[ValidatedSatellite, str, str]] = []
    for satellite in cluster.satellites:
        if satellite.cardinality == "one_to_one":
            for header in satellite.payload_columns:
                columns.append((satellite, header, "scalar"))
        else:
            for header in satellite.payload_columns:
                columns.append((satellite, header, "summary"))
    return columns


def _chunks_from_enriched_primary_sheet(
    sheet: SheetData,
    *,
    cluster: ValidatedCluster,
    lookups: dict[str, dict[str, Any]],
) -> list[ExtractedChunk]:
    if not sheet.rows:
        return []

    band_size = resolve_row_band_size(len(sheet.headers))
    col_range = [1, len(sheet.headers)]
    chunks: list[ExtractedChunk] = []

    enrichment_columns = _enrichment_columns_for_cluster(cluster)
    enriched_headers = list(sheet.headers)
    for satellite, payload_header, mode in enrichment_columns:
        if mode == "scalar":
            enriched_headers.append(_field_label(satellite, payload_header))
        else:
            enriched_headers.append(_field_label(satellite, f"{payload_header}_summary"))

    for band_index, band_start in enumerate(range(0, len(sheet.rows), band_size)):
        band_rows = sheet.rows[band_start : band_start + band_size]
        data_rows: list[list[str]] = []
        sheet_row_map: list[int] = []
        entity_keys: list[str] = []
        row_entity_keys: dict[str, str] = {}
        enrichment_origins: list[dict[str, Any]] = []

        for row_num, cells in band_rows:
            extra_values, origins, entity_key = _enrich_primary_row(
                sheet,
                row_num,
                cells,
                cluster=cluster,
                lookups=lookups,
                enrichment_columns=enrichment_columns,
            )
            if entity_key:
                entity_keys.append(entity_key)
                row_entity_keys[str(row_num)] = entity_key
            enrichment_origins.extend(origins)
            padded_cells = list(cells) + [""] * max(0, len(sheet.headers) - len(cells))
            row_values = padded_cells[: len(sheet.headers)] + extra_values
            data_rows.append(row_values)

        content = rows_to_slim_values_text(data_rows)
        if not content or len(content.split()) < MIN_TABLE_WORDS:
            continue

        sheet_row_map = [row_num for row_num, _ in band_rows]
        extra: dict[str, Any] = {
            "source_format": "xlsx",
            "sheet_index": sheet.index,
            "sheet_name": sheet.name,
            "table_index": band_index,
            "row_range": [sheet_row_map[0], sheet_row_map[-1]],
            "col_range": col_range,
            "section": sheet.name,
            "table_headers": enriched_headers,
            "sheet_row_map": sheet_row_map,
            "content_format": "slim_rows",
            "sheet_role": "primary",
            "cluster_primary_sheet": cluster.primary_sheet,
            "entity_key_column": cluster.primary_key_column,
            "entity_keys": sorted(set(entity_keys)),
            "row_entity_keys": row_entity_keys,
            "enrichment_origins": enrichment_origins,
        }

        chunks.append(
            ExtractedChunk(
                content=content,
                page_number=sheet.index,
                chunk_type="table",
                extraction_method="xlsx_native",
                extra_metadata=extra,
            )
        )
    return chunks


def _chunks_from_satellite_sheet(
    sheet: SheetData,
    *,
    cluster: ValidatedCluster,
    satellite: ValidatedSatellite,
) -> list[ExtractedChunk]:
    if satellite.cardinality != "one_to_many":
        return []

    band_size = resolve_row_band_size(len(sheet.headers))
    col_range = [1, len(sheet.headers)]
    chunks: list[ExtractedChunk] = []

    for band_index, band_start in enumerate(range(0, len(sheet.rows), band_size)):
        band_rows = sheet.rows[band_start : band_start + band_size]
        data_rows = [cells for _, cells in band_rows]
        content = rows_to_slim_values_text(data_rows)
        if not content or len(content.split()) < MIN_TABLE_WORDS:
            continue

        sheet_row_map = [row_num for row_num, _ in band_rows]
        entity_keys: list[str] = []
        row_entity_keys: dict[str, str] = {}
        for row_num, cells in band_rows:
            if satellite.key_col_index < len(cells):
                key_value = cells[satellite.key_col_index].strip()
                if key_value:
                    entity_keys.append(key_value)
                    row_entity_keys[str(row_num)] = key_value

        extra: dict[str, Any] = {
            "source_format": "xlsx",
            "sheet_index": sheet.index,
            "sheet_name": sheet.name,
            "table_index": band_index,
            "row_range": [sheet_row_map[0], sheet_row_map[-1]],
            "col_range": col_range,
            "section": sheet.name,
            "table_headers": list(sheet.headers),
            "sheet_row_map": sheet_row_map,
            "content_format": "slim_rows",
            "sheet_role": "satellite",
            "cluster_primary_sheet": cluster.primary_sheet,
            "entity_key_column": satellite.key_column,
            "entity_keys": sorted(set(entity_keys)),
            "row_entity_keys": row_entity_keys,
            "satellite_cardinality": satellite.cardinality,
        }

        chunks.append(
            ExtractedChunk(
                content=content,
                page_number=sheet.index,
                chunk_type="table",
                extraction_method="xlsx_native",
                extra_metadata=extra,
            )
        )
    return chunks


def _annotate_standalone_chunks_for_foreign_keys(
    chunks: list[ExtractedChunk],
    sheet: SheetData,
    schema: ValidatedWorkbookSchema,
) -> None:
    for cluster in schema.clusters:
        key_index = resolve_header_index(sheet.headers, cluster.primary_key_column)
        if key_index is None:
            continue
        for chunk in chunks:
            annotate_chunk_entity_keys(
                chunk,
                key_col_index=key_index,
                key_column=cluster.primary_key_column,
            )
        return


def _chunks_from_openpyxl_sheet(xlsx_path: str, sheet: SheetData) -> list[ExtractedChunk]:
    """Fallback to native table/band extraction for standalone sheets."""
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path, data_only=True, read_only=False)
    try:
        worksheet = workbook[sheet.name]
        table_chunks = _chunks_from_tables(worksheet, sheet.index, sheet.name)
        if table_chunks:
            for chunk in table_chunks:
                chunk.extra_metadata["sheet_role"] = "standalone"
            return table_chunks
        chunks = _chunks_from_row_bands(worksheet, sheet.index, sheet.name)
        for chunk in chunks:
            chunk.extra_metadata["sheet_role"] = "standalone"
        return chunks
    finally:
        workbook.close()


def build_workbook_chunks(
    xlsx_path: str,
    workbook: WorkbookData,
    schema: ValidatedWorkbookSchema,
    *,
    doc_id: str | None = None,
    user_id: int | None = None,
) -> list[ExtractedChunk]:
    chunks: list[ExtractedChunk] = []
    handled_sheets: set[str] = set()

    for cluster in schema.clusters:
        primary_sheet = workbook.sheet_by_name.get(cluster.primary_sheet)
        if primary_sheet is None:
            continue
        lookups = _build_satellite_lookups(workbook, cluster)
        chunks.extend(
            _chunks_from_enriched_primary_sheet(
                primary_sheet,
                cluster=cluster,
                lookups=lookups,
            )
        )
        handled_sheets.add(cluster.primary_sheet)

        for satellite in cluster.satellites:
            satellite_sheet = workbook.sheet_by_name.get(satellite.sheet_name)
            if satellite_sheet is None:
                continue
            if satellite.cardinality == "one_to_many":
                chunks.extend(
                    _chunks_from_satellite_sheet(
                        satellite_sheet,
                        cluster=cluster,
                        satellite=satellite,
                    )
                )
            handled_sheets.add(satellite.sheet_name)

    standalone_names = set(schema.standalone_sheets) | {
        sheet.name for sheet in workbook.sheets if sheet.name not in handled_sheets
    }
    for sheet_name in sorted(standalone_names):
        sheet = workbook.sheet_by_name.get(sheet_name)
        if sheet is None:
            continue
        sheet_chunks = _chunks_from_openpyxl_sheet(xlsx_path, sheet)
        _annotate_standalone_chunks_for_foreign_keys(sheet_chunks, sheet, schema)
        chunks.extend(sheet_chunks)
        handled_sheets.add(sheet_name)

    logger.info(
        "XLSX_SCHEMA_ENRICH doc_id=%s user_id=%s clusters=%s standalone=%s chunks=%s",
        doc_id,
        user_id,
        len(schema.clusters),
        len(standalone_names),
        len(chunks),
    )
    return chunks


def extract_xlsx_workbook(
    xlsx_path: str,
    *,
    doc_id: str | None = None,
    user_id: int | None = None,
) -> tuple[list[ExtractedChunk], ValidatedWorkbookSchema]:
    """Full XLSX pipeline: load workbook, detect schema, build enriched chunks.

    Schema detection (cross-sheet join enrichment) is best-effort: it calls an
    LLM, which can fail transiently (rate limits, timeouts, outages). That
    failure must not sink ingestion of the underlying table data, so on
    failure every sheet falls back to being treated as standalone -- the
    same outcome as a sheet the LLM legitimately found no joins for.
    """
    workbook = load_workbook_data(xlsx_path)
    try:
        schema = detect_and_validate_workbook_schema(workbook)
    except WorkbookSchemaRecognitionError as exc:
        logger.warning(
            "XLSX_SCHEMA_FALLBACK doc_id=%s user_id=%s reason=%s — proceeding with all sheets standalone",
            doc_id,
            user_id,
            exc.detail,
        )
        schema = ValidatedWorkbookSchema(llm_error=exc.detail)
    chunks = build_workbook_chunks(
        xlsx_path,
        workbook,
        schema,
        doc_id=doc_id,
        user_id=user_id,
    )
    return chunks, schema
