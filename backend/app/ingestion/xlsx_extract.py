"""XLSX ingestion: sheet-aware table extraction.

Hybrid chunking: Excel Table objects (ListObjects) when present; otherwise row-banded
used ranges per sheet. Hidden and very-hidden sheets are skipped. Cell values are
read with ``data_only=True`` so formulas contribute their last-computed values.

Row-banded sheets index compact pipe-delimited value rows with column headers stored
once in chunk metadata (not repeated per band).
"""

from __future__ import annotations

import logging
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.ingestion.chunking import normalize_whitespace
from app.ingestion.models import ExtractedChunk
from app.ingestion.tables import table_signature, table_to_markdown
from app.ingestion.xlsx_serialize import resolve_row_band_size, rows_to_slim_values_text

logger = logging.getLogger(__name__)

MIN_TABLE_WORDS = 4


def is_sheet_visible(worksheet: Worksheet) -> bool:
    state = getattr(worksheet, "sheet_state", "visible") or "visible"
    return state == "visible"


def _cell_value(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return normalize_whitespace(value)
    return normalize_whitespace(str(value))


def is_title_row(cells: list[str]) -> bool:
    """A row with at most one non-empty cell reads as a banner/title (e.g.
    'FINANCIAL HIGHLIGHTS' alone in column A), not real column headers --
    common above the header row in report-style exports."""
    return sum(1 for cell in cells if cell) <= 1


def find_header_index(numbered: list[tuple[int, list[str]]]) -> int:
    """Index into `numbered` of the first row that looks like real headers,
    skipping any leading title/banner rows. Falls back to 0 if every row
    looks like a title (nothing better to pick)."""
    for index, (_, cells) in enumerate(numbered):
        if not is_title_row(cells):
            return index
    return 0


# Report-style tables sometimes split the header across two physical rows,
# e.g. row 4 = "2024 | | 2025 |" (group labels, sparse) and row 5 =
# "Actual | Budget | Actual | Budget" (dense sub-labels) before the real
# numeric data starts on row 6. A second header row is distinguished from
# the first real data row by comparing against a sample of rows further into
# the table: real data rows in a genuinely tabular sheet are mostly numeric,
# while a header continuation row is mostly text even when the surrounding
# data is numeric. This self-calibrates instead of using an absolute
# threshold, so it doesn't misfire on sheets that are naturally text-heavy
# throughout (e.g. a title/description catalog) -- there, the "body" sample
# is text-heavy too, so a text-heavy row 2 no longer looks anomalous.
_HEADER_BODY_SAMPLE_ROWS = 5
_HEADER_BODY_MIN_NUMERIC_FRACTION = 0.5
_HEADER_CONTINUATION_MAX_NUMERIC_FRACTION = 0.2

_NUMERIC_LIKE_RE = re.compile(r"^\(?-?[\d,]+\.?\d*\)?%?$")


def _looks_numeric(value: str) -> bool:
    return bool(value) and bool(_NUMERIC_LIKE_RE.match(value.strip()))


def _numeric_fraction(cells: list[str]) -> float:
    non_empty = [cell for cell in cells if cell]
    if not non_empty:
        return 0.0
    return sum(1 for cell in non_empty if _looks_numeric(cell)) / len(non_empty)


def _fill_ratio(cells: list[str]) -> float:
    if not cells:
        return 0.0
    return sum(1 for cell in cells if cell) / len(cells)


def find_header_span(numbered: list[tuple[int, list[str]]]) -> tuple[int, int]:
    """Like `find_header_index`, but returns a (start, end) half-open range
    into `numbered` covering one row normally, or two when the row right
    after the header looks like a second header level rather than data."""
    start = find_header_index(numbered)
    end = start + 1

    look_ahead = numbered[end + 1 : end + 1 + _HEADER_BODY_SAMPLE_ROWS]
    if not look_ahead:
        return start, end

    body_numeric = sum(_numeric_fraction(cells) for _, cells in look_ahead) / len(look_ahead)
    if body_numeric < _HEADER_BODY_MIN_NUMERIC_FRACTION:
        return start, end

    next_row = numbered[end][1]
    if _numeric_fraction(next_row) > _HEADER_CONTINUATION_MAX_NUMERIC_FRACTION:
        return start, end
    if _fill_ratio(next_row) < _fill_ratio(numbered[start][1]):
        return start, end

    return start, end + 1


def _forward_fill(row: list[str]) -> list[str]:
    """Propagate each non-empty cell rightward until the next non-empty one.
    A merged group-label cell (e.g. "2024" spanning two columns) only has a
    value in its leftmost column when read via openpyxl; siblings read as
    empty. This reconstructs the intended per-column group label."""
    filled: list[str] = []
    last = ""
    for cell in row:
        if cell.strip():
            last = cell.strip()
        filled.append(last)
    return filled


def resolve_header_merges(
    worksheet: Worksheet,
    header_rows: list[tuple[int, list[str]]],
    min_col: int,
) -> list[list[str]]:
    """Fill blank header cells that are genuinely part of an Excel merged
    range (e.g. "2024" merged across two columns, so only the leftmost
    column holds the value and its sibling reads as blank) with the merge's
    anchor value. Uses real merge metadata rather than guessing from
    blankness alone -- unlike `_forward_fill`, this never risks filling a
    column that's simply blank for unrelated reasons, since it only acts on
    cells Excel itself records as merged. Applies to a single header row
    too, which `merge_header_rows`'s forward-fill deliberately does not."""
    resolved = [list(cells) for _, cells in header_rows]
    row_numbers = [row_num for row_num, _ in header_rows]
    for merged_range in worksheet.merged_cells.ranges:
        overlapping = [i for i, row_num in enumerate(row_numbers) if merged_range.min_row <= row_num <= merged_range.max_row]
        if not overlapping:
            continue
        anchor_value = _cell_value(worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value)
        if not anchor_value:
            continue
        for i in overlapping:
            row_cells = resolved[i]
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                index = col - min_col
                if 0 <= index < len(row_cells) and not row_cells[index].strip():
                    row_cells[index] = anchor_value
    return resolved


def merge_header_rows(rows: list[list[str]]) -> list[str]:
    """Combine one or more header rows into a single header list, e.g.
    ["", "2024", "2025"] + ["Metric", "Actual", "Actual"] -> ["Metric",
    "2024 Actual", "2025 Actual"]. A single row passes through unchanged
    (no forward-fill -- that's only meaningful when reconstructing a group
    label shared across rows, not within one already-complete header row)."""
    if len(rows) == 1:
        return list(rows[0])
    filled_rows = [_forward_fill(row) for row in rows]
    width = max(len(row) for row in filled_rows)
    merged: list[str] = []
    for col in range(width):
        parts = []
        for row in filled_rows:
            if col < len(row) and row[col] and row[col] not in parts:
                parts.append(row[col])
        merged.append(" ".join(parts))
    return merged


def rows_from_range(
    worksheet: Worksheet,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
) -> list[list[str]]:
    return [cells for _, cells in rows_from_range_numbered(worksheet, min_row, min_col, max_row, max_col)]


def rows_from_range_numbered(
    worksheet: Worksheet,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
) -> list[tuple[int, list[str]]]:
    rows: list[tuple[int, list[str]]] = []
    for sheet_row, row in enumerate(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ),
        start=min_row,
    ):
        cells = [_cell_value(cell) for cell in row]
        if any(cells):
            rows.append((sheet_row, cells))
    return rows


def _base_metadata(
    sheet_index: int,
    sheet_name: str,
    *,
    table_index: int,
    row_range: list[int],
    col_range: list[int],
) -> dict[str, Any]:
    return {
        "source_format": "xlsx",
        "sheet_index": sheet_index,
        "sheet_name": sheet_name,
        "table_index": table_index,
        "row_range": row_range,
        "col_range": col_range,
        "section": sheet_name,
    }


def _table_chunk(
    rows: list[list[str]],
    sheet_index: int,
    sheet_name: str,
    table_index: int,
    row_range: list[int],
    col_range: list[int],
    *,
    sheet_row_map: list[int],
) -> ExtractedChunk | None:
    markdown = table_to_markdown(rows)
    if not markdown or len(markdown.split()) < MIN_TABLE_WORDS:
        return None

    extra = _base_metadata(
        sheet_index,
        sheet_name,
        table_index=table_index,
        row_range=row_range,
        col_range=col_range,
    )
    _, headers = table_signature(rows)
    if headers:
        extra["table_headers"] = list(headers)
    extra["sheet_row_map"] = sheet_row_map
    extra["content_format"] = "markdown_table"

    return ExtractedChunk(
        content=markdown,
        page_number=sheet_index,
        chunk_type="table",
        extraction_method="xlsx_native",
        extra_metadata=extra,
    )


def _band_chunk(
    header: list[str],
    band_numbered: list[tuple[int, list[str]]],
    sheet_index: int,
    sheet_name: str,
    table_index: int,
    col_range: list[int],
) -> ExtractedChunk | None:
    data_rows = [cells for _, cells in band_numbered]
    content = rows_to_slim_values_text(data_rows)
    if not content or len(content.split()) < MIN_TABLE_WORDS:
        return None

    sheet_row_map = [sheet_row for sheet_row, _ in band_numbered]
    extra = _base_metadata(
        sheet_index,
        sheet_name,
        table_index=table_index,
        row_range=[sheet_row_map[0], sheet_row_map[-1]],
        col_range=col_range,
    )
    extra["table_headers"] = list(header)
    extra["sheet_row_map"] = sheet_row_map
    extra["content_format"] = "slim_rows"

    return ExtractedChunk(
        content=content,
        page_number=sheet_index,
        chunk_type="table",
        extraction_method="xlsx_native",
        extra_metadata=extra,
    )


def _chunks_from_tables(worksheet: Worksheet, sheet_index: int, sheet_name: str) -> list[ExtractedChunk]:
    chunks: list[ExtractedChunk] = []
    tables = list(worksheet.tables.values())
    for table_index, table in enumerate(tables):
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        numbered = rows_from_range_numbered(worksheet, min_row, min_col, max_row, max_col)
        rows = [cells for _, cells in numbered]
        sheet_row_map = [sheet_row for sheet_row, _ in numbered]
        chunk = _table_chunk(
            rows,
            sheet_index,
            sheet_name,
            table_index,
            row_range=[sheet_row_map[0], sheet_row_map[-1]] if sheet_row_map else [min_row, max_row],
            col_range=[min_col, max_col],
            sheet_row_map=sheet_row_map,
        )
        if chunk is not None:
            chunks.append(chunk)
    return chunks


def _chunks_from_row_bands(worksheet: Worksheet, sheet_index: int, sheet_name: str) -> list[ExtractedChunk]:
    if worksheet.max_row is None or worksheet.max_column is None:
        return []
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return []

    column_count = worksheet.max_column
    band_size = resolve_row_band_size(column_count)
    col_range = [1, column_count]

    all_numbered = rows_from_range_numbered(worksheet, 1, 1, worksheet.max_row, worksheet.max_column)
    if len(all_numbered) < 2:
        sheet_row_map = [sheet_row for sheet_row, _ in all_numbered]
        chunk = _table_chunk(
            [cells for _, cells in all_numbered],
            sheet_index,
            sheet_name,
            table_index=0,
            row_range=[sheet_row_map[0], sheet_row_map[-1]] if sheet_row_map else [1, worksheet.max_row],
            col_range=col_range,
            sheet_row_map=sheet_row_map,
        )
        return [chunk] if chunk is not None else []

    header_start, header_end = find_header_span(all_numbered)
    header_rows = resolve_header_merges(worksheet, all_numbered[header_start:header_end], min_col=1)
    header = merge_header_rows(header_rows)
    data_numbered = all_numbered[header_end:]
    chunks: list[ExtractedChunk] = []

    for band_index, band_start in enumerate(range(0, len(data_numbered), band_size)):
        band_numbered = data_numbered[band_start : band_start + band_size]
        chunk = _band_chunk(
            header,
            band_numbered,
            sheet_index,
            sheet_name,
            table_index=band_index,
            col_range=col_range,
        )
        if chunk is not None:
            chunks.append(chunk)

    logger.info(
        "Row-banded sheet %r cols=%s band_size=%s chunks=%s",
        sheet_name,
        column_count,
        band_size,
        len(chunks),
    )
    return chunks


def extract_xlsx_chunks(
    xlsx_path: str,
    *,
    doc_id: str | None = None,
    user_id: int | None = None,
) -> list[ExtractedChunk]:
    """Extract table chunks from an XLSX workbook."""
    workbook = load_workbook(xlsx_path, data_only=True, read_only=False)
    chunks: list[ExtractedChunk] = []
    visible_sheet_index = 0

    try:
        for worksheet in workbook.worksheets:
            if not is_sheet_visible(worksheet):
                continue
            visible_sheet_index += 1
            sheet_name = worksheet.title

            table_chunks = _chunks_from_tables(worksheet, visible_sheet_index, sheet_name)
            if table_chunks:
                chunks.extend(table_chunks)
            else:
                chunks.extend(_chunks_from_row_bands(worksheet, visible_sheet_index, sheet_name))
    finally:
        workbook.close()

    logger.info(
        "Extracted %s chunks from XLSX doc_id=%s user_id=%s",
        len(chunks),
        doc_id,
        user_id,
    )
    return chunks


def count_visible_sheets(xlsx_path: str) -> int:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        return sum(1 for ws in workbook.worksheets if is_sheet_visible(ws))
    finally:
        workbook.close()
