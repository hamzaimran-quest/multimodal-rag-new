"""XLSX embedded picture extraction (OCR-proxy indexing).

Mirrors the DOCX image pipeline in ``docx_images.py``: raster pictures
embedded in a sheet (openpyxl ``worksheet._images``) are saved to disk, OCR'd,
and indexed as ``chunk_type="image"`` chunks with nearby cell text as proxy
content. Like every other image path in this system, these chunks are
retrievable via hybrid search but are never fed to the answer LLM -- no
vision-model reasoning happens anywhere here, by design.

Native Excel chart objects (``worksheet._charts``) are out of scope: their
underlying numbers already exist as real cell values elsewhere in the sheet
and get table-chunked normally, so the chart image itself would mostly be
redundant with data already indexed.
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import settings
from app.ingestion.chunking import normalize_whitespace
from app.ingestion.images import NEARBY_TEXT_WORD_LIMIT, _build_image_chunk_content, _ocr_image_text
from app.ingestion.models import ExtractedChunk
from app.ingestion.xlsx_extract import is_sheet_visible

logger = logging.getLogger(__name__)

MIN_IMAGE_PIXELS = 2_500
# Anchor is the image's top-left cell; pictures are usually placed above/left
# of the content they illustrate and extend down/right from there, so the
# look-ahead window is asymmetric rather than a plain square radius.
ANCHOR_ROWS_BEFORE = 2
ANCHOR_ROWS_AFTER = 10
ANCHOR_COLS_BEFORE = 2
ANCHOR_COLS_AFTER = 10
# Sheets at or under this size use their *entire* content as nearby text
# instead of an anchor-radius window. Small "cover"/"example"/"instructions"
# sheets that exist mainly to host a picture often have all their sparse
# text far from the anchor (e.g. a single title cell in row 1, image anchored
# several rows down) -- a local radius would miss it, and there's little risk
# of pulling in unrelated content when the whole sheet is this small anyway.
SMALL_SHEET_ROW_THRESHOLD = 20
SMALL_SHEET_COL_THRESHOLD = 20


def _save_image_blob(blob: bytes, image_path: Path) -> bool:
    try:
        from PIL import Image
    except Exception:
        return False
    try:
        with Image.open(io.BytesIO(blob)) as img:
            width, height = img.size
            if width * height < MIN_IMAGE_PIXELS:
                return False
            if img.mode not in ("RGB", "RGBA"):
                img = img.convert("RGB")
            img.save(image_path, format="PNG")
        return True
    except Exception:
        logger.debug("Failed to persist XLSX image %s", image_path, exc_info=True)
        return False


def _anchor_row_col(image: Any) -> tuple[int, int] | None:
    """0-indexed (row, col) of an embedded image's anchor cell, or None if
    the anchor shape isn't one openpyxl exposes a cell position for."""
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None) if anchor is not None else None
    if marker is None or not hasattr(marker, "row") or not hasattr(marker, "col"):
        return None
    return int(marker.row), int(marker.col)


def _nearby_text(worksheet: Worksheet, anchor_row_1idx: int, anchor_col_1idx: int) -> str:
    # worksheet.max_row/max_column only reflect cell content, not embedded
    # drawings -- a sheet can be almost empty of cells (or even entirely
    # empty) and still host a large picture, e.g. an "example"/template
    # sheet whose only content is an explanatory diagram. Use whichever is
    # larger so the window always reaches at least the anchor itself.
    real_max_row = worksheet.max_row or 0
    real_max_col = worksheet.max_column or 0
    sheet_max_row = max(real_max_row, anchor_row_1idx)
    sheet_max_col = max(real_max_col, anchor_col_1idx)

    if real_max_row <= SMALL_SHEET_ROW_THRESHOLD and real_max_col <= SMALL_SHEET_COL_THRESHOLD:
        min_row, max_row = 1, sheet_max_row
        min_col, max_col = 1, sheet_max_col
    else:
        min_row = max(1, anchor_row_1idx - ANCHOR_ROWS_BEFORE)
        max_row = min(sheet_max_row, anchor_row_1idx + ANCHOR_ROWS_AFTER)
        min_col = max(1, anchor_col_1idx - ANCHOR_COLS_BEFORE)
        max_col = min(sheet_max_col, anchor_col_1idx + ANCHOR_COLS_AFTER)
    if min_row > max_row or min_col > max_col:
        return ""

    parts: list[str] = []
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
        for value in row:
            if value is None:
                continue
            text = normalize_whitespace(str(value)).strip()
            if text:
                parts.append(text)
    if not parts:
        return ""
    return normalize_whitespace(" ".join(parts))[:NEARBY_TEXT_WORD_LIMIT]


def extract_xlsx_image_chunks(
    xlsx_path: str,
    *,
    doc_id: str | None = None,
    user_id: int | None = None,
) -> list[ExtractedChunk]:
    """Extract embedded pictures from every visible sheet as OCR-proxy image chunks."""
    if not doc_id or user_id is None:
        return []

    workbook = load_workbook(xlsx_path, data_only=True, read_only=False)
    out_dir = settings.resolved_images_dir / str(user_id) / doc_id
    chunks: list[ExtractedChunk] = []

    try:
        visible_index = 0
        for worksheet in workbook.worksheets:
            if not is_sheet_visible(worksheet):
                continue
            visible_index += 1
            images = list(getattr(worksheet, "_images", []) or [])
            if not images:
                continue

            out_dir.mkdir(parents=True, exist_ok=True)
            for img_idx, image in enumerate(images):
                position = _anchor_row_col(image)
                anchor_row = position[0] + 1 if position else None
                anchor_col = position[1] + 1 if position else None

                try:
                    blob = image.ref.getvalue() if hasattr(image.ref, "getvalue") else image._data()
                except Exception:
                    logger.debug("Failed to read XLSX image blob sheet=%s index=%s", worksheet.title, img_idx, exc_info=True)
                    continue

                image_filename = f"sheet{visible_index}_img{img_idx}.png"
                image_path = out_dir / image_filename
                if not _save_image_blob(blob, image_path):
                    continue

                nearby = _nearby_text(worksheet, anchor_row, anchor_col) if anchor_row and anchor_col else ""
                ocr_text = _ocr_image_text(image_path)
                if not nearby and not ocr_text:
                    image_path.unlink(missing_ok=True)
                    continue

                content = _build_image_chunk_content(
                    nearby_text=nearby,
                    ocr_text=ocr_text,
                    page_number=visible_index,
                )
                relative_image_path = f"data/images/{user_id}/{doc_id}/{image_filename}"
                extra: dict[str, Any] = {
                    "source_format": "xlsx",
                    "sheet_index": visible_index,
                    "sheet_name": worksheet.title,
                    "section": worksheet.title,
                    "anchor_row": anchor_row,
                    "anchor_col": anchor_col,
                    "image_caption": nearby[:200] if nearby else "",
                    "ocr_available": bool(ocr_text),
                }
                chunks.append(
                    ExtractedChunk(
                        content=content,
                        page_number=visible_index,
                        chunk_type="image",
                        extraction_method="xlsx_ocr_proxy",
                        image_path=relative_image_path,
                        extra_metadata=extra,
                    )
                )
    finally:
        workbook.close()

    logger.info("Extracted %s image chunks from XLSX doc_id=%s user_id=%s", len(chunks), doc_id, user_id)
    return chunks
