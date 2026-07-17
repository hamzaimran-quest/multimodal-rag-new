"""Load XLSX workbook structure for schema detection and enrichment."""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook

from app.ingestion.xlsx_extract import (
    find_header_span,
    is_sheet_visible,
    merge_header_rows,
    resolve_header_merges,
    rows_from_range_numbered,
)


def normalize_header(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_cell(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def headers_match(left: str, right: str) -> bool:
    return normalize_header(left).casefold() == normalize_header(right).casefold()


def resolve_header_index(headers: list[str], column_name: str) -> int | None:
    target = normalize_header(column_name).casefold()
    for index, header in enumerate(headers):
        if normalize_header(header).casefold() == target:
            return index
    return None


@dataclass
class SheetData:
    name: str
    index: int
    headers: list[str]
    rows: list[tuple[int, list[str]]]
    header_to_col: dict[str, int] = field(default_factory=dict)

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def column_count(self) -> int:
        return len(self.headers)

    def sample_rows(self, limit: int) -> list[tuple[int, list[str]]]:
        return self.rows[:limit]

    def column_values(self, column_name: str, *, normalized: bool = True) -> list[str]:
        index = resolve_header_index(self.headers, column_name)
        if index is None:
            return []
        values: list[str] = []
        for _, cells in self.rows:
            if index >= len(cells):
                continue
            raw = cells[index]
            values.append(normalize_cell(raw) if normalized else raw)
        return values

    def cell_at(self, sheet_row: int, column_name: str) -> str:
        index = resolve_header_index(self.headers, column_name)
        if index is None:
            return ""
        for row_num, cells in self.rows:
            if row_num == sheet_row and index < len(cells):
                return cells[index]
        return ""


@dataclass
class WorkbookData:
    sheets: list[SheetData]
    sheet_by_name: dict[str, SheetData] = field(default_factory=dict)

    def __post_init__(self) -> None:
        if not self.sheet_by_name:
            self.sheet_by_name = {sheet.name: sheet for sheet in self.sheets}


def load_workbook_data(xlsx_path: str) -> WorkbookData:
    """Read all visible sheets. The first row that isn't a blank/title banner
    is treated as the header, extended with a second row when that row looks
    like a header continuation (e.g. group labels split across two rows)
    rather than data -- see `find_header_span`. Everything after is data."""
    workbook = load_workbook(xlsx_path, data_only=True, read_only=False)
    sheets: list[SheetData] = []
    visible_index = 0
    try:
        for worksheet in workbook.worksheets:
            if not is_sheet_visible(worksheet):
                continue
            visible_index += 1
            max_row = int(worksheet.max_row or 0)
            max_col = int(worksheet.max_column or 0)
            if max_row < 1 or max_col < 1:
                continue

            numbered = rows_from_range_numbered(worksheet, 1, 1, max_row, max_col)
            if not numbered:
                continue

            header_start, header_end = find_header_span(numbered)
            header_rows = resolve_header_merges(worksheet, numbered[header_start:header_end], min_col=1)
            header_cells = merge_header_rows(header_rows)
            headers = [normalize_header(cell) for cell in header_cells]
            data_rows = [(row_num, cells) for row_num, cells in numbered[header_end:]]

            header_to_col = {
                normalize_header(header).casefold(): col_index
                for col_index, header in enumerate(headers)
                if normalize_header(header)
            }
            sheets.append(
                SheetData(
                    name=worksheet.title,
                    index=visible_index,
                    headers=headers,
                    rows=data_rows,
                    header_to_col=header_to_col,
                )
            )
    finally:
        workbook.close()
    return WorkbookData(sheets=sheets)


def build_workbook_summary_for_llm(
    workbook: WorkbookData,
    *,
    sample_row_limit: int,
) -> dict[str, Any]:
    """Compact workbook description for one-shot schema LLM call."""
    sheet_summaries: list[dict[str, Any]] = []
    for sheet in workbook.sheets:
        samples: list[dict[str, Any]] = []
        for row_num, cells in sheet.sample_rows(sample_row_limit):
            row_dict = {
                sheet.headers[col_index]: cells[col_index]
                for col_index in range(min(len(sheet.headers), len(cells)))
                if normalize_header(sheet.headers[col_index])
            }
            samples.append({"row": row_num, "values": row_dict})
        sheet_summaries.append(
            {
                "name": sheet.name,
                "row_count": sheet.row_count,
                "column_count": sheet.column_count,
                "headers": sheet.headers,
                "sample_rows": samples,
            }
        )
    return {"sheets": sheet_summaries}
