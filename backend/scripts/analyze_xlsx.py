"""Quick XLSX structure + chunk size analysis."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.ingestion.xlsx_extract import count_visible_sheets, extract_xlsx_chunks


def main(path: str) -> None:
    wb = load_workbook(path, data_only=True, read_only=False)
    print("=== WORKBOOK OVERVIEW ===")
    print(f"Sheets: {len(wb.sheetnames)}")
    for i, name in enumerate(wb.sheetnames):
        ws = wb[name]
        state = getattr(ws, "sheet_state", "visible")
        max_r, max_c = ws.max_row, ws.max_column
        tables = list(ws.tables.values())
        print(f"  [{i+1}] {name!r} state={state} rows={max_r} cols={max_c} tables={len(tables)}")
        for t in tables:
            mc, mr, xc, xr = range_boundaries(t.ref)
            print(f"      table {t.displayName!r} ref={t.ref} ({xr-mr+1} rows x {xc-mc+1} cols)")
    wb.close()

    chunks = extract_xlsx_chunks(path, doc_id="test", user_id=1)
    print(f"\n=== CHUNKS: {len(chunks)} visible_sheets={count_visible_sheets(path)} ===")
    total_chars = 0
    total_words = 0
    for c in chunks:
        extra = c.extra_metadata or {}
        words = len(c.content.split())
        chars = len(c.content)
        lines = c.content.count("\n") + 1
        total_chars += chars
        total_words += words
        est_tokens = chars // 4
        print(
            f"  sheet={extra.get('sheet_name')} words={words} chars={chars} "
            f"est_tokens~{est_tokens} lines={lines} "
            f"row_range={extra.get('row_range')} col_range={extra.get('col_range')}"
        )
        preview = c.content[:250].replace("\n", " | ")
        print(f"    preview: {preview}...")
    print(f"\nTOTAL: {total_words} words, {total_chars} chars, est_tokens~{total_chars//4}")


if __name__ == "__main__":
    main(sys.argv[1])
