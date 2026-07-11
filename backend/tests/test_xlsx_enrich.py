"""Tests for enriched XLSX chunk building."""

from __future__ import annotations

from pathlib import Path

from app.ingestion.xlsx_enrich import build_workbook_chunks
from app.ingestion.xlsx_schema import (
    ValidatedCluster,
    ValidatedSatellite,
    ValidatedWorkbookSchema,
)
from app.ingestion.xlsx_workbook import load_workbook_data
from tests.xlsx_fixtures import build_star_schema_xlsx


def test_enriched_primary_includes_scalar_and_cast_summary(tmp_path: Path) -> None:
    xlsx_path = build_star_schema_xlsx(tmp_path / "star.xlsx")
    workbook = load_workbook_data(str(xlsx_path))
    schema = ValidatedWorkbookSchema(
        clusters=[
            ValidatedCluster(
                primary_sheet="Shows",
                primary_key_column="show_id",
                primary_key_col_index=0,
                satellites=[
                    ValidatedSatellite(
                        sheet_name="Countries",
                        key_column="show_id",
                        key_col_index=0,
                        payload_columns=["country"],
                        payload_col_indices=[1],
                        cardinality="one_to_one",
                        overlap_ratio=1.0,
                        key_repeat_ratio=0.0,
                    ),
                    ValidatedSatellite(
                        sheet_name="Cast",
                        key_column="show_id",
                        key_col_index=0,
                        payload_columns=["actor"],
                        payload_col_indices=[1],
                        cardinality="one_to_many",
                        overlap_ratio=1.0,
                        key_repeat_ratio=0.5,
                    ),
                ],
            )
        ],
        standalone_sheets=[],
    )

    chunks = build_workbook_chunks(str(xlsx_path), workbook, schema)
    primary_chunks = [
        chunk for chunk in chunks if chunk.extra_metadata.get("sheet_role") == "primary"
    ]
    satellite_chunks = [
        chunk for chunk in chunks if chunk.extra_metadata.get("sheet_role") == "satellite"
    ]

    assert primary_chunks
    assert satellite_chunks
    primary = primary_chunks[0]
    assert "Countries.country" in primary.extra_metadata["table_headers"]
    assert "Cast.actor_summary" in primary.extra_metadata["table_headers"]
    assert "France" in primary.content
    assert "Alice" in primary.content
    assert primary.extra_metadata["entity_keys"] == ["1", "2"]
    assert primary.extra_metadata["row_entity_keys"] == {"2": "1", "3": "2"}
    assert primary.extra_metadata.get("enrichment_origins")

    satellite = satellite_chunks[0]
    assert "Alice" in satellite.content
    assert satellite.extra_metadata["entity_keys"]


def test_standalone_sheet_gets_foreign_key_metadata(tmp_path: Path) -> None:
    from openpyxl import Workbook

    from app.ingestion.xlsx_enrich import build_workbook_chunks
    from app.ingestion.xlsx_schema import ValidatedCluster, ValidatedWorkbookSchema
    from app.ingestion.xlsx_workbook import load_workbook_data

    xlsx_path = tmp_path / "standalone_fk.xlsx"
    workbook = Workbook()
    shows = workbook.active
    shows.title = "Shows"
    shows.append(["show_id", "title"])
    shows.append(["1", "Alpha"])

    countries = workbook.create_sheet("Countries")
    countries.append(["show_id", "country"])
    countries.append(["1", "France"])
    countries.append(["2", "Germany"])
    countries.append(["3", "Spain"])
    countries.append(["4", "Italy"])
    workbook.save(xlsx_path)
    workbook.close()

    loaded = load_workbook_data(str(xlsx_path))
    schema = ValidatedWorkbookSchema(
        clusters=[
            ValidatedCluster(
                primary_sheet="Shows",
                primary_key_column="show_id",
                primary_key_col_index=0,
                satellites=[],
            )
        ],
        standalone_sheets=["Countries"],
    )
    chunks = build_workbook_chunks(str(xlsx_path), loaded, schema)
    country_chunks = [
        chunk for chunk in chunks if chunk.extra_metadata.get("sheet_name") == "Countries"
    ]
    assert country_chunks
    assert country_chunks[0].extra_metadata["entity_keys"] == ["1", "2", "3", "4"]
    assert country_chunks[0].extra_metadata["row_entity_keys"]["2"] == "1"
