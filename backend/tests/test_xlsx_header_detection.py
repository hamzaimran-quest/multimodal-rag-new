"""Tests for header-row detection: title/banner-row skipping and two-row
(hierarchical) header merging, at both the low-level row-scanning functions
and through the full load_workbook_data / extract_xlsx_chunks pipelines."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.ingestion.xlsx_extract import (
    extract_xlsx_chunks,
    find_header_span,
    merge_header_rows,
    resolve_header_merges,
    rows_from_range_numbered,
)
from app.ingestion.xlsx_workbook import load_workbook_data
from tests.xlsx_fixtures import (
    build_merged_header_xlsx,
    build_text_heavy_single_header_xlsx,
    build_title_banner_xlsx,
    build_two_row_header_xlsx,
)


def _numbered_rows(xlsx_path: Path, sheet_name: str):
    workbook = load_workbook(str(xlsx_path), data_only=True)
    try:
        worksheet = workbook[sheet_name]
        return rows_from_range_numbered(worksheet, 1, 1, worksheet.max_row, worksheet.max_column)
    finally:
        workbook.close()


def test_find_header_span_skips_title_banner_and_uses_single_row(tmp_path: Path):
    xlsx_path = build_title_banner_xlsx(tmp_path / "banner.xlsx")
    numbered = _numbered_rows(xlsx_path, "Financial highlights")

    start, end = find_header_span(numbered)

    # Blank row 1 is already filtered out by rows_from_range_numbered, so the
    # title row is numbered[0]; the real header ("CHF million", ...) is next.
    assert numbered[start][1][0] == "CHF million"
    assert end == start + 1  # single header row, no false-positive merge

    header = merge_header_rows([cells for _, cells in numbered[start:end]])
    assert header == ["CHF million", "2024", "2025"]


def test_load_workbook_data_skips_title_banner(tmp_path: Path):
    xlsx_path = build_title_banner_xlsx(tmp_path / "banner.xlsx")
    workbook = load_workbook_data(str(xlsx_path))

    assert len(workbook.sheets) == 1
    sheet = workbook.sheets[0]
    assert sheet.headers == ["CHF million", "2024", "2025"]
    assert sheet.row_count == 2
    assert sheet.rows[0][1] == ["Order intake", "725.5", "703.4"]


def test_find_header_span_merges_genuine_two_row_header(tmp_path: Path):
    xlsx_path = build_two_row_header_xlsx(tmp_path / "two_row.xlsx")
    numbered = _numbered_rows(xlsx_path, "Budget")

    start, end = find_header_span(numbered)

    assert end == start + 2  # both header rows consumed
    header = merge_header_rows([cells for _, cells in numbered[start:end]])
    assert header == ["Metric", "2024 Actual", "2024 Budget", "2025 Actual", "2025 Budget"]

    # First row after the merged header must be real numeric data, not the
    # "Actual/Budget" sub-header row.
    first_data_row = numbered[end][1]
    assert first_data_row[0] == "Revenue"
    assert first_data_row[1] == "100"


def test_load_workbook_data_merges_two_row_header(tmp_path: Path):
    xlsx_path = build_two_row_header_xlsx(tmp_path / "two_row.xlsx")
    workbook = load_workbook_data(str(xlsx_path))

    sheet = workbook.sheets[0]
    assert sheet.headers == ["Metric", "2024 Actual", "2024 Budget", "2025 Actual", "2025 Budget"]
    assert sheet.row_count == 3
    assert sheet.rows[0][1][0] == "Revenue"


def test_find_header_span_does_not_swallow_text_heavy_data_row(tmp_path: Path):
    """Regression guard: a naturally text-heavy single-header table (titles,
    descriptions) must not have its first data row misread as a second
    header level just because it's mostly non-numeric."""
    xlsx_path = build_text_heavy_single_header_xlsx(tmp_path / "titles.xlsx")
    numbered = _numbered_rows(xlsx_path, "Titles")

    start, end = find_header_span(numbered)

    assert end == start + 1
    header = merge_header_rows([cells for _, cells in numbered[start:end]])
    assert header == ["duration_minutes", "type", "title", "rating", "description"]

    first_data_row = numbered[end][1]
    assert first_data_row[2] == "Norm of the North"


def test_extract_xlsx_chunks_uses_merged_two_row_header(tmp_path: Path):
    xlsx_path = build_two_row_header_xlsx(tmp_path / "two_row.xlsx")
    chunks = extract_xlsx_chunks(str(xlsx_path), doc_id="doc-2row", user_id=1)

    assert len(chunks) == 1
    chunk = chunks[0]
    assert chunk.extra_metadata["table_headers"] == [
        "Metric",
        "2024 Actual",
        "2024 Budget",
        "2025 Actual",
        "2025 Budget",
    ]
    assert "Revenue | 100 | 110 | 120 | 115" in chunk.content
    # The "Actual/Budget" sub-header row must not leak into the data content.
    assert "Actual | Budget" not in chunk.content


def test_resolve_header_merges_fills_merged_cells_but_not_unrelated_blanks(tmp_path: Path):
    xlsx_path = build_merged_header_xlsx(tmp_path / "merged.xlsx")
    workbook = load_workbook(str(xlsx_path), data_only=True)
    try:
        worksheet = workbook["Highlights"]
        numbered = rows_from_range_numbered(worksheet, 1, 1, worksheet.max_row, worksheet.max_column)
        start, end = find_header_span(numbered)
        assert end == start + 1  # single physical header row

        resolved = resolve_header_merges(worksheet, numbered[start:end], min_col=1)
        header = merge_header_rows(resolved)
    finally:
        workbook.close()

    # B1:C1 merged as "2024", D1:E1 merged as "2025" -> both siblings filled.
    assert header == ["CHF million", "2024", "2024", "2025", "2025", ""]


def test_load_workbook_data_resolves_merged_header_cells(tmp_path: Path):
    xlsx_path = build_merged_header_xlsx(tmp_path / "merged.xlsx")
    workbook = load_workbook_data(str(xlsx_path))

    sheet = workbook.sheets[0]
    assert sheet.headers == ["CHF million", "2024", "2024", "2025", "2025", ""]
    # column_values still resolves the (now-filled) header name to the right column.
    assert sheet.column_values("2025") == ["703.4", "859.1"]


def test_extract_xlsx_chunks_resolves_merged_header_cells(tmp_path: Path):
    xlsx_path = build_merged_header_xlsx(tmp_path / "merged.xlsx")
    chunks = extract_xlsx_chunks(str(xlsx_path), doc_id="doc-merged", user_id=1)

    assert len(chunks) == 1
    assert chunks[0].extra_metadata["table_headers"] == ["CHF million", "2024", "2024", "2025", "2025", ""]
