"""Unit tests for XLSX embedded picture extraction (OCR-proxy indexing)."""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

from app.config import settings
from app.ingestion.xlsx_images import MIN_IMAGE_PIXELS, extract_xlsx_image_chunks


def _png_bytes(size: tuple[int, int] = (120, 120), color: tuple[int, int, int] = (200, 40, 40)) -> io.BytesIO:
    img = Image.new("RGB", size, color=color)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


def _write_xlsx_with_anchored_image(path: Path, *, anchor: str = "B2", hidden: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"
    for row in range(1, 8):
        sheet.append([f"Metric {row}", row * 10, row * 20])
    sheet.add_image(XLImage(_png_bytes()), anchor)
    if hidden:
        # A workbook can't hide its only sheet -- add a second, visible one.
        workbook.create_sheet("Other")
        sheet.sheet_state = "hidden"
    workbook.save(path)
    workbook.close()


def test_extract_xlsx_image_chunks_persists_and_indexes(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "sample.xlsx"
    _write_xlsx_with_anchored_image(xlsx_path, anchor="B2")

    chunks = extract_xlsx_image_chunks(str(xlsx_path), doc_id="doc-1", user_id=7)

    assert len(chunks) == 1
    chunk = chunks[0]
    assert chunk.chunk_type == "image"
    assert chunk.extraction_method == "xlsx_ocr_proxy"
    assert chunk.page_number == 1
    assert chunk.extra_metadata["source_format"] == "xlsx"
    assert chunk.extra_metadata["sheet_name"] == "Report"
    assert chunk.extra_metadata["anchor_row"] == 2
    assert chunk.extra_metadata["anchor_col"] == 2
    assert chunk.bbox is None
    assert "Metric" in chunk.content

    image_file = settings.resolved_images_dir / "7" / "doc-1" / "sheet1_img0.png"
    assert image_file.is_file()


def test_extract_xlsx_image_chunks_skips_without_ids(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "sample.xlsx"
    _write_xlsx_with_anchored_image(xlsx_path)
    assert extract_xlsx_image_chunks(str(xlsx_path)) == []


def test_extract_xlsx_image_chunks_skips_hidden_sheet(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "sample.xlsx"
    _write_xlsx_with_anchored_image(xlsx_path, hidden=True)
    assert extract_xlsx_image_chunks(str(xlsx_path), doc_id="doc-1", user_id=7) == []


def test_extract_xlsx_image_chunks_skips_too_small_image(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Metric", "Value"])
    tiny_side = int(MIN_IMAGE_PIXELS**0.5) - 5
    sheet.add_image(XLImage(_png_bytes(size=(tiny_side, tiny_side))), "A2")
    xlsx_path = tmp_path / "tiny.xlsx"
    workbook.save(xlsx_path)
    workbook.close()

    assert extract_xlsx_image_chunks(str(xlsx_path), doc_id="doc-1", user_id=7) == []


def test_extract_xlsx_image_chunks_uses_whole_sheet_for_sparse_cover_sheet(tmp_path: Path) -> None:
    """Regression test: a near-empty "cover"/"example" sheet whose only cell
    content sits far from the image anchor (title in row 1, image anchored
    several rows down) must still pick up that title as nearby text -- a
    fixed local radius window would miss it entirely, exactly what happened
    with the real Finances.xlsx "Example Cashflow Statement" sheet this was
    built against (single title cell in row 1, image anchored at row 4)."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Example Cashflow Statement"
    sheet["A1"] = "Example cashflow statement"
    sheet.add_image(XLImage(_png_bytes()), "B4")
    xlsx_path = tmp_path / "sparse.xlsx"
    workbook.save(xlsx_path)
    workbook.close()

    chunks = extract_xlsx_image_chunks(str(xlsx_path), doc_id="doc-2", user_id=7)

    assert len(chunks) == 1
    assert "Example cashflow statement" in chunks[0].content
