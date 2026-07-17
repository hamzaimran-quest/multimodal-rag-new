"""Helpers for building sample XLSX files in tests."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table


def build_star_schema_xlsx(path: Path) -> Path:
    """Workbook with one primary sheet and 1:1 + 1:many satellites for schema tests."""
    workbook = Workbook()
    shows = workbook.active
    shows.title = "Shows"
    shows.append(["show_id", "title", "year"])
    shows.append(["1", "Alpha", "2020"])
    shows.append(["2", "Beta", "2021"])

    countries = workbook.create_sheet("Countries")
    countries.append(["show_id", "country"])
    countries.append(["1", "France"])
    countries.append(["2", "Germany"])

    cast = workbook.create_sheet("Cast")
    cast.append(["show_id", "actor"])
    cast.append(["1", "Alice"])
    cast.append(["1", "Bob"])
    cast.append(["2", "Carol"])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def build_title_banner_xlsx(path: Path) -> Path:
    """Report-style sheet: blank row 1, single-cell title row 2, blank row 3,
    real headers on row 4 -- regression fixture for the title/banner-row
    header-detection fix (see find_header_index / is_title_row)."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Financial highlights"
    sheet.append([None, None, None])
    sheet.append(["FINANCIAL HIGHLIGHTS", None, None])
    sheet.append([None, None, None])
    sheet.append(["CHF million", "2024", "2025"])
    sheet.append(["Order intake", 725.5, 703.4])
    sheet.append(["Sales", 969.2, 859.1])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def build_two_row_header_xlsx(path: Path) -> Path:
    """Report-style sheet with a header split across two physical rows: row 1
    has sparse group labels ("2024"/"2025"), row 2 has dense sub-labels
    ("Actual"/"Budget"), and only then does numeric data start on row 3."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Budget"
    sheet.append([None, "2024", None, "2025", None])
    sheet.append(["Metric", "Actual", "Budget", "Actual", "Budget"])
    sheet.append(["Revenue", 100, 110, 120, 115])
    sheet.append(["Costs", 60, 65, 70, 68])
    sheet.append(["Profit", 40, 45, 50, 47])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def build_text_heavy_single_header_xlsx(path: Path) -> Path:
    """Single-row-header sheet whose data rows are mostly text (titles,
    descriptions) with only a couple of numeric columns -- false-positive
    guard for two-row header detection, modeled on catalog-style data
    (e.g. netflix_titles) where naive "row 2 looks non-numeric" logic would
    wrongly swallow the first real data row into the header."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Titles"
    sheet.append(["duration_minutes", "type", "title", "rating", "description"])
    sheet.append([90, "Movie", "Norm of the North", "TV-PG", "A polar bear king must take back a stolen artifact."])
    sheet.append([94, "Movie", "Jandino: Whatever it Takes", "TV-MA", "Jandino riffs on raising kids."])
    sheet.append([45, "TV Show", "Transformers Prime", "TV-Y7", "Autobots protect Earth from the Decepticons."])
    sheet.append([50, "TV Show", "Transformers: Robots in Disguise", "TV-Y7", "Bumblebee leads a new Autobot force."])
    sheet.append([60, "Movie", "Whatever It Takes", "TV-14", "A drama about family and ambition."])
    sheet.append([70, "Movie", "Something Else", "TV-G", "A documentary about nothing in particular."])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def build_merged_header_xlsx(path: Path) -> Path:
    """Single-row header where "2024" and "2025" are each a real Excel merge
    spanning 2 columns (e.g. an "Actual" / "Budget" pair under each year),
    plus a genuinely blank-headered but data-bearing column at the end --
    that last one must stay blank (unmerged), not get filled in by the
    merge resolution just because its header cell happens to be empty."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Highlights"
    sheet.append(["CHF million", "2024", None, "2025", None, None])
    sheet.merge_cells("B1:C1")
    sheet.merge_cells("D1:E1")
    sheet.append(["Order intake", 725.5, 700.0, 703.4, 690.0, "note-a"])
    sheet.append(["Sales", 969.2, 950.0, 859.1, 840.0, "note-b"])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def build_sample_xlsx(path: Path) -> Path:
    workbook = Workbook()
    revenue = workbook.active
    revenue.title = "Revenue"
    revenue.append(["Year", "Revenue"])
    revenue.append(["2023", 100])
    revenue.append(["2024", 120])
    revenue.add_table(Table(displayName="RevenueTable", ref="A1:B3"))

    hidden = workbook.create_sheet("Hidden")
    hidden.append(["Secret", "Value"])
    hidden.append(["X", 1])
    hidden.sheet_state = "hidden"

    bands = workbook.create_sheet("Bands")
    bands.append(["Metric", "Value"])
    for index in range(1, 6):
        bands.append([f"Row {index}", index * 10])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path
